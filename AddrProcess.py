# -*- coding: utf-8 -*-
__author__ = 'few'
# 创建时间 2018/3/4 16:39 
import xlrd,xlwt,sys
from Logger import logger
import openpyxl
openpyxl.load_workbook('Data.xlsx')
class AddrProcess():
    def __init__(self):
        self.addr_manage = {}

    def addrInit(self):
        try:
            data = xlrd.open_workbook('Data.xlsx')#打开xls文件
        except IOError as err:
            logger.error(err,exc_info=True)#使用参数exc_info=True调用logger方法，trackback会输出到logger中。
            sys.exit(0)
        else:
            table = data.sheet_by_name(u'Address')#选择 Logger 数据页
            if table.nrows>1:
                for i in range(1,table.nrows):
                    if table.row_values(i)[0] !=  table.row_values(i)[1] !=  table.row_values(i)[2] !=  '':
                        self.addr_manage[str(table.row_values(i)[0])] = [(str(table.row_values(i)[1]),int(table.row_values(i)[2])),i+1]
                        # self.addr_manage[str(table.row_values(i)[0])] = ('192.168.0.1',9999)

            logger.info(self.addr_manage)
            return self.addr_manage

    def writeExcel(self,row,column,value):
        workbook = openpyxl.load_workbook('Data.xlsx')
        table = workbook.get_sheet_by_name('Address')
        table.cell(row=row, column=column, value=value)
        # table['A4'] = 1
        # table.cell(row=4, column=2, value=10)
        workbook.save('Data.xlsx')
if __name__ == "__main__":
    ap = AddrProcess()
    ap.addrInit()
